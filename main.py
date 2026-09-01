print("1. 🚀 El script de Python HA INICIADO su ejecución...") # <-- Rastreo 1

import os
import re
import secrets
import string
import smtplib
import math
import io
import openpyxl
import locale
import openpyxl.drawing.image as openpyxl_image
from werkzeug.utils import secure_filename
from datetime import datetime
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.application import MIMEApplication
import pandas as pd
from flask import Flask, render_template, request, redirect, url_for, flash, session, jsonify, send_file
from database.models import Colaborador, CategoriaDemografica, ValorDemografico, Participacion, Base, Empresa, Usuario
from sqlalchemy import create_engine, func, text
from sqlalchemy.orm import sessionmaker
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

# Librerías oficiales de Google API
from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials
from googleapiclient.discovery import build

print("2. 📚 Librerías importadas correctamente...") # <-- Rastreo 2

app = Flask(__name__)
app.secret_key = 'clave_secreta_super_segura_para_sesiones'

# Configuración de carpeta de subidas de logos
UPLOAD_FOLDER = os.path.join(app.root_path, 'static', 'logos_empresas')
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg'}

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

@app.route('/admin/cargar-logo-empresa', methods=['POST'])
def cargar_logo_empresa():
    if 'usuario_id' not in session or session['rol'] not in ['admin', 'coordinador']:
        return redirect('/login')

    empresa_id = request.form.get('empresa_id')
    archivo = request.files.get('logo_cliente')

    if not empresa_id or not archivo or archivo.filename == '':
        flash("⚠️ Por favor seleccione la empresa y una imagen válida.", "warning")
        return redirect('/admin')

    if archivo and allowed_file(archivo.filename):
        db = SessionLocal()
        try:
            filename = secure_filename(f"logo_empresa_{empresa_id}_{archivo.filename}")
            filepath = os.path.join(UPLOAD_FOLDER, filename)
            archivo.save(filepath)

            # Guardar la ruta relativa en Neon
            tabla_real = Empresa.__table__.name
            db.execute(
                text(f"UPDATE {tabla_real} SET logo_url = :url WHERE id = :id"),
                {"url": f"/static/logos_empresas/{filename}", "id": int(empresa_id)}
            )
            db.commit()
            flash("🎨 Logo de la organización actualizado con éxito.", "success")
        except Exception as e:
            db.rollback()
            flash(f"❌ Error al guardar la imagen: {str(e)}", "danger")
        finally:
            db.close()
    else:
        flash("⚠️ Formato de imagen no permitido (Use PNG, JPG o JPEG).", "warning")

    return redirect('/admin')

# =================================================================
# 📅 MOTOR 1: INTEGRACIÓN CON GOOGLE CALENDAR API (CREAR Y BORRAR)
# =================================================================
def crear_evento_google_calendar(nombre_empresa, fecha_inicio_str, fecha_cierre_str, correo_coordinador):
    SCOPES = ['https://www.googleapis.com/auth/calendar']
    creds = None
    
    if os.path.exists('token.json'):
        creds = Credentials.from_authorized_user_file('token.json', SCOPES)
        
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            try:
                creds.refresh(Request())
            except Exception as e:
                print(f"❌ Error al refrescar token de Google: {e}")
                return None
        else:
            print("⚠️ Advertencia: Falta token.json corporativo.")
            return None

    try:
        service = build('calendar', 'v3', credentials=creds)
        fecha_cierre_limpia = fecha_cierre_str.replace("-", "").replace("/", "")
        regra_recurrencia = f"RRULE:FREQ=DAILY;UNTIL={fecha_cierre_limpia}T235959Z"
        
        evento = {
            'summary': f'🔄 Actualizar reporte HR™ - {nombre_empresa}',
            'description': f'Tarea automática: Recuerde cargar el reporte actualizado de {nombre_empresa}.',
            'start': {
                'dateTime': f'{fecha_inicio_str}T08:00:00',
                'timeZone': 'America/Bogota',
            },
            'end': {
                'dateTime': f'{fecha_inicio_str}T08:01:00',
                'timeZone': 'America/Bogota',
            },
            'recurrence': [regra_recurrencia],
            'attendees': [{'email': correo_coordinador}],
            'reminders': {
                'useDefault': False,
                'overrides': [{'method': 'popup', 'minutes': 5}],
            },
        }

        evento_creado = service.events().insert(calendarId='primary', body=evento, sendUpdates='all').execute()
        print(f"📅 [CALENDAR] Tarea agendada correctamente.")
        return evento_creado.get('id')
    except Exception as e:
        print(f"❌ ERROR AL CREAR EVENTO EN CALENDAR: {e}")
        return None

def eliminar_evento_google_calendar(event_id):
    SCOPES = ['https://www.googleapis.com/auth/calendar']
    if not event_id or event_id == '-': return False
    
    if os.path.exists('token.json'):
        creds = Credentials.from_authorized_user_file('token.json', SCOPES)
        if creds and creds.valid:
            try:
                service = build('calendar', 'v3', credentials=creds)
                service.events().delete(calendarId='primary', eventId=event_id, sendUpdates='all').execute()
                print(f"[CALENDAR] Tarea recurrente {event_id} eliminada con éxito de Google.")
                return True
            except Exception as e:
                print(f"❌ Error al intentar remover el evento de Google Calendar: {e}")
    return False

# =================================================================
# 📧 MOTOR 2: ENVÍO REAL DE CORREOS MEDIANTE GMAIL SMTP
# =================================================================
def generar_password_aleatorio(longitud=10):
    caracteres = string.ascii_letters + string.digits
    return ''.join(secrets.choice(caracteres) for _ in range(longitud))

def generar_username_cliente(nombre_completo):
    limpio = nombre_completo.lower().strip().replace('á','a').replace('é','e').replace('í','i').replace('ó','o').replace('ú','u').replace('ñ','n')
    partes = limpio.split()
    if len(partes) >= 2:
        username = partes[0][0] + partes[1] + "01"
    elif len(partes) == 1:
        username = partes[0] + "01"
    else:
        username = "cliente01"
    return re.sub(r'[^a-z0-9]', '', username)

def enviar_correo_notificacion(destinatario, nombre_titular, usuario_login, password_login, es_olvido=False):
    remitente = "carlos.mora@peoplesvoice.co"
    alias_marca = "Portal de Reportes HR™"
    
    msg = MIMEMultipart()
    msg['From'] = f"{alias_marca} <{remitente}>"
    msg['To'] = destinatario
    
    if es_olvido:
        msg['Subject'] = "Restablecimiento de Contraseña - Portal de Reportes HR™"
        html = f"""
        <html>
        <body style="font-family: 'Helvetica Neue', Helvetica, Arial, sans-serif; color: #333; line-height: 1.6;">
            <div style="max-width: 550px; margin: 0 auto; padding: 25px; border: 1px solid #edf2f7; border-radius: 10px; background-color: #ffffff;">
                <h2 style="color: #3498db; font-size: 1.4rem; margin-bottom: 15px; font-weight: 700;">Nueva Contraseña Asignada</h2>
                <p>Hola <strong>{nombre_titular}</strong>,</p>
                <p>Hemos procesado tu solicitud de restablecimiento de contraseña corporativa:</p>
                <div style="background-color: #f8f9fa; padding: 15px; border-radius: 6px; margin: 20px 0; border-left: 4px solid #3498db;">
                    <p style="margin: 0 0 8px 0;"><strong>Usuario de Acceso:</strong> <code style="font-size: 1rem; color: #2c3e50; font-weight: bold;">{usuario_login}</code></p>
                    <p style="margin: 0;"><strong>Nueva Contraseña:</strong> <code style="font-size: 1rem; color: #e74c3c; font-weight: bold;">{password_login}</code></p>
                </div>
                <hr style="border: 0; border-top: 1px solid #edf2f7; margin: 20px 0;">
                <p style="font-size: 0.78rem; color: #a0aec0; text-align: center;">Portal de Reportes HR™ — Peoples Voice © 2026</p>
            </div>
        </body>
        </html>
        """
    else:
        msg['Subject'] = "Bienvenido al Portal de Reportes HR™ - Credenciales de Acceso"
        html = f"""
        <html>
        <body style="font-family: 'Helvetica Neue', Helvetica, Arial, sans-serif; color: #333; line-height: 1.6;">
            <div style="max-width: 550px; margin: 0 auto; padding: 25px; border: 1px solid #edf2f7; border-radius: 10px; background-color: #ffffff;">
                <h2 style="color: #2ecc71; font-size: 1.4rem; margin-bottom: 15px; font-weight: 700;">¡Bienvenido a la Plataforma!</h2>
                <p>Hola <strong>{nombre_titular}</strong>,</p>
                <p>Se ha cargado exitosamente tu organización en nuestro system; desde ahora puedes acceder a los reportes de participación de la encuesta de HR™.
                Tus credenciales de ingreso oficiales son:</p>
                <div style="background-color: #f8f9fa; padding: 15px; border-radius: 6px; margin: 20px 0; border-left: 4px solid #2ecc71;">
                    <p style="margin: 0 0 8px 0;"><strong>Usuario Asigned:</strong> <code style="font-size: 1rem; color: #2c3e50; font-weight: bold;">{usuario_login}</code></p>
                    <p style="margin: 0;"><strong>Contraseña del Sistema:</strong> <code style="font-size: 1rem; color: #3498db; font-weight: bold;">{password_login}</code></p>
                </div>
                <hr style="border: 0; border-top: 1px solid #edf2f7; margin: 20px 0;">
                <p style="font-size: 0.78rem; color: #a0aec0; text-align: center;">Portal de Reportes HR™ — Peoples Voice © 2026</p>
            </div>
        </body>
        </html>
        """
    
    msg.attach(MIMEText(html, 'html'))
    
    SMTP_SERVER = "smtp.gmail.com"
    SMTP_PORT = 587
    SMTP_PASSWORD = "hxhjkhqleflgvmoo" 
    
    try:
        server = smtplib.SMTP(SMTP_SERVER, SMTP_PORT)
        server.starttls()
        server.login(remitente, SMTP_PASSWORD) 
        server.sendmail(remitente, destinatario, msg.as_string()) 
        server.quit()
        print(f"📧 [SMTP GMAIL] Correo real enviado con éxito a {destinatario}")
        return True
    except Exception as e:
        print(f"❌ ERROR SMTP GMAIL: {e}")
        return False

# =================================================================
# CONFIGURACIÓN E INICIALIZACIÓN DE LA BASE DE DATOS
# =================================================================
try:
    DATABASE_URL = os.environ.get(
        "DATABASE_URL",
        "postgresql://neondb_owner:npg_DvFw5XoZ9Yuh@ep-fancy-feather-aws87y8x-pooler.c-12.us-east-1.aws.neon.tech/neondb?sslmode=require&channel_binding=require"
    )
    
    engine = create_engine(
        DATABASE_URL,
        pool_size=5,
        max_overflow=10,
        pool_recycle=30,
        pool_pre_ping=True
    )
    SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)
    
    print("3. 🗄️ Intentando conectar a Neon y crear las tablas automáticamente...") 
    Base.metadata.create_all(bind=engine)
    
    tabla_real = Empresa.__table__.name
    tabla_usuario = Usuario.__table__.name
    
    conn = engine.connect()
    try:
        for col, col_type in [("fecha_inicio", "TEXT"), ("fecha_cierre", "TEXT"), ("cerrada", "INTEGER DEFAULT 0"), ("nombre_cliente", "TEXT"), ("correo_cliente", "TEXT"), ("correo_coordinador", "TEXT"), ("calendar_event_id", "TEXT")]:
            try:
                conn.execute(text(f"ALTER TABLE {tabla_real} ADD COLUMN {col} {col_type}"))
                conn.connect().commit() if hasattr(conn, 'commit') else conn.commit()
            except Exception:
                pass
                
        try:
            conn.execute(text(f"ALTER TABLE {tabla_usuario} ADD COLUMN nombre TEXT"))
            conn.connect().commit() if hasattr(conn, 'commit') else conn.commit()
        except Exception:
            pass
    finally:
        conn.close() 
    
    db_init = SessionLocal()
    try:
        user_admin_existe = db_init.query(Usuario).filter(Usuario.email == "carlos.mora@peoplesvoice.co").first()
        if not user_admin_existe:
            nuevo_admin = Usuario(email="carlos.mora@peoplesvoice.co", password_hash="Colombia2026*", rol="admin", empresa_id=None)
            db_init.add(nuevo_admin)
            db_init.flush()
            db_init.execute(text(f"UPDATE {tabla_usuario} SET nombre = 'Carlos Mora' WHERE email = 'carlos.mora@peoplesvoice.co'"))
            db_init.commit()
        else:
            db_init.execute(text(f"UPDATE {tabla_usuario} SET nombre = 'Carlos Mora', password_hash = 'Colombia2026*' WHERE email = 'carlos.mora@peoplesvoice.co'"))
            db_init.commit()
    finally:
        db_init.close() 
        
    print("4. ✅ Tablas y usuarios verificados sin bloqueos.") 
except Exception as e:
    print(f"❌ ERROR EN BASE DE DATOS: {e}")

# ==========================================
# RUTAS DE CONTROL GENERAL Y AUTENTICACIÓN
# ==========================================
@app.route('/')
def inicio():
    if 'usuario_id' not in session: return redirect('/login')
    return redirect('/admin')

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        email = request.form.get('email').strip().lower()
        password = request.form.get('password')
        
        db = SessionLocal()
        usuario = db.query(Usuario).filter(Usuario.email == email).first()
        db.close()
        
        if usuario and usuario.password_hash == password: 
            session['usuario_id'] = usuario.id
            session['email'] = usuario.email
            session['rol'] = usuario.rol
            session['empresa_id'] = usuario.empresa_id
            session['nombre'] = getattr(usuario, 'nombre', usuario.email) or usuario.email
            
            return redirect('/admin')
            
        flash('Credenciales incorrectas', 'danger')
    return render_template('login.html')

@app.route('/admin')
def panel_admin():
    if 'usuario_id' not in session or session['rol'] not in ['admin', 'coordinador']: 
        return redirect('/login')
        
    db = SessionLocal()
    organizaciones_maestras = []
    coordinadores_maestros = []
    
    try:
        tabla_real = Empresa.__table__.name
        resultado_raw = db.execute(text(f"SELECT id, nombre, fecha_inicio, fecha_cierre, cerrada, nombre_cliente, correo_cliente, correo_coordinador, calendar_event_id FROM {tabla_real}")).fetchall()
        
        for fila in resultado_raw:
            usr = db.query(Usuario).filter(Usuario.empresa_id == fila[0], Usuario.rol == 'cliente').first()
            
            f_ini = fila[2] if len(fila) > 2 and fila[2] else '-'
            f_cie = fila[3] if len(fila) > 3 and fila[3] else '-'
            cer = 1 if len(fila) > 4 and fila[4] == 1 else 0
            n_cli = fila[5] if len(fila) > 5 and fila[5] else '-'
            c_cli = fila[6] if len(fila) > 6 and fila[6] else '-'
            c_coor = fila[7] if len(fila) > 7 and fila[7] else '-'
            ev_id = fila[8] if len(fila) > 8 and fila[8] else '-'
            
            organizaciones_maestras.append({
                "id": fila[0],
                "nombre": fila[1],
                "fecha_inicio": f_ini,
                "fecha_cierre": f_cie,
                "cerrada": cer,
                "nombre_cliente": n_cli,
                "correo_cliente": c_cli,
                "correo_coordinador": c_coor,
                "calendar_event_id": ev_id,
                "email": usr.email if usr else "Sin canal",
                "password": usr.password_hash if usr else "Sin clave"
            })
            
        if session.get('rol') == 'admin':
            lista_c = db.query(Usuario).filter(Usuario.rol == 'coordinador').all()
            for c in lista_c:
                coordinadores_maestros.append({
                    "id": c.id,
                    "nombre": getattr(c, 'nombre', c.email) or c.email,
                    "email": c.email,
                    "password": c.password_hash
                })
    except Exception as e:
        print(f"❌ Error al cargar panel admin: {e}")
        flash(f"Error al cargar datos del panel: {str(e)}", "danger")
    finally:
        db.close()
        
    return render_template('admin.html', empresas=organizaciones_maestras, organizaciones=organizaciones_maestras, coordinadores=coordinadores_maestros)

@app.route('/admin/crear-coordinador', methods=['POST'])
def crear_coordinador():
    if 'usuario_id' not in session or session['rol'] != 'admin': return redirect('/login')
    nombre_coor = request.form.get('nombre_coordinador', '').strip()
    email_coor = request.form.get('email_coordinador', '').strip().lower()
    
    if nombre_coor and email_coor:
        db = SessionLocal()
        existe = db.query(Usuario).filter(Usuario.email == email_coor).first()
        if not existe:
            pass_coor = generar_password_aleatorio()
            nuevo_c = Usuario(email=email_coor, password_hash=pass_coor, rol='coordinador', empresa_id=None)
            db.add(nuevo_c)
            db.flush()
            
            tabla_usuario = Usuario.__table__.name
            db.execute(text(f"UPDATE {tabla_usuario} SET nombre = :nom WHERE id = :id"), {"nom": nombre_coor, "id": nuevo_c.id})
            db.commit()
            flash(f"👥 Coordinador '{nombre_coor}' registrado. Contraseña asignada: {pass_coor}", "success")
        else:
            flash("⚠️ Ese correo electrónico ya está asignado.", "danger")
        db.close()
    return redirect('/admin')

@app.route('/admin/cambiar-mi-password', methods=['POST'])
def cambiar_mi_password():
    if 'usuario_id' not in session: return redirect('/login')
    nueva_pass = request.form.get('nueva_password', '').strip()
    
    if nueva_pass:
        db = SessionLocal()
        usuario = db.query(Usuario).filter(Usuario.id == session['usuario_id']).first()
        if usuario:
            usuario.password_hash = nueva_pass
            db.commit()
            flash("🔒 Contraseña actualizada con éxito.", "success")
        db.close()
    return redirect('/admin')

@app.route('/admin/crear-empresa', methods=['POST'])
def crear_empresa():
    if 'usuario_id' not in session or session['rol'] not in ['admin', 'coordinador']: 
        return redirect('/login')
        
    nombre_emp = (request.form.get('organizacion') or request.form.get('nombre_empresa') or '').strip()
    correo_coor = (request.form.get('correo_cdp') or request.form.get('correo_coordinador') or '').strip().lower()
    fecha_ini = request.form.get('fecha_inicio', '').strip()
    fecha_cie = request.form.get('fecha_cierre', '').strip()
    
    if nombre_emp:
        db = SessionLocal()
        try:
            if not db.query(Empresa).filter(Empresa.nombre == nombre_emp).first():
                nueva_empresa = Empresa(nombre=nombre_emp)
                db.add(nueva_empresa)
                db.flush()
                
                event_id = '-'
                if fecha_ini and fecha_cie and correo_coor:
                    try:
                        event_id = crear_evento_google_calendar(nombre_emp, fecha_ini, fecha_cie, correo_coor) or '-'
                    except Exception as e:
                        print(f"❌ Alerta Calendar: {e}")
                        event_id = '-'
                
                tabla_real = Empresa.__table__.name
                db.execute(
                    text(f"UPDATE {tabla_real} SET fecha_inicio=:ini, fecha_cierre=:cie, cerrada=0, correo_coordinador=:c_coor, calendar_event_id=:ev_id WHERE id=:id"),
                    {"ini": fecha_ini, "cie": fecha_cie, "c_coor": correo_coor, "ev_id": event_id, "id": nueva_empresa.id}
                )
                
                db.commit()
                flash(f"🏢 ¡Organización '{nombre_emp}' creada con éxito!", "success")
            else:
                flash("⚠️ Esa Organización ya se encuentra registrada.", "danger")
        except Exception as e:
            db.rollback()
            flash(f"❌ Error en base de datos: {str(e)}", "danger")
        finally:
            db.close()
    else:
        flash("⚠️ Por favor ingrese el nombre de la Organización.", "warning")
        
    return redirect('/admin')

@app.route('/admin/editar-empresa/<int:empresa_id>', methods=['POST'])
def editar_empresa(empresa_id):
    if 'usuario_id' not in session or session['rol'] not in ['admin', 'coordinador']: return redirect('/login')
    db = SessionLocal()
    nuevo_nombre = request.form.get('nombre_empresa', '').strip()
    fecha_ini = request.form.get('fecha_inicio', '').strip()
    fecha_cie = request.form.get('fecha_cierre', '').strip()
    nombre_cli = request.form.get('nombre_cliente', '').strip()
    correo_cli = request.form.get('correo_cliente', '').strip().lower()
    correo_coor = request.form.get('correo_coordinador', '').strip().lower()
    
    if nuevo_nombre:
        tabla_real = Empresa.__table__.name
        db.execute(text(f"UPDATE {tabla_real} SET nombre=:nom, fecha_inicio=:ini, fecha_cierre=:cie, nombre_cliente=:n_cli, correo_cliente=:c_cli, correo_coordinador=:c_coor WHERE id=:id"),
                {"nom": nuevo_nombre, "ini": fecha_ini, "cie": fecha_cie, "n_cli": nombre_cli, "c_cli": correo_cli, "c_coor": correo_coor, "id": empresa_id})
        db.commit()
        flash("✏️ ¡Cambios guardados correctamente!", "success")
    db.close()
    return redirect('/admin')

@app.route('/admin/eliminar-organizacion/<int:empresa_id>')
def eliminar_organizacion(empresa_id):
    if 'usuario_id' not in session or session['rol'] not in ['admin', 'coordinador']: return redirect('/login')
    db = SessionLocal()
    try:
        tabla_real = Empresa.__table__.name
        res_ev = db.execute(text(f"SELECT calendar_event_id FROM {tabla_real} WHERE id = :id"), {"id": empresa_id}).fetchone()
        if res_ev and res_ev[0] and res_ev[0] != '-':
            eliminar_evento_google_calendar(res_ev[0])
            
        emp_id_int = int(empresa_id)
        colab_ids = [c[0] for c in db.query(Colaborador.id).filter(Colaborador.empresa_id == emp_id_int).all()]
        if colab_ids:
            db.execute(text("DELETE FROM valores_demograficos WHERE colaborador_id = ANY(:ids)"), {"ids": colab_ids})
            db.execute(text("DELETE FROM participaciones WHERE colaborador_id = ANY(:ids)"), {"ids": colab_ids})
            db.execute(text("DELETE FROM colaboradores WHERE empresa_id = :emp_id"), {"emp_id": emp_id_int})
            
        db.execute(text("DELETE FROM categorias_demograficas WHERE empresa_id = :emp_id"), {"emp_id": emp_id_int})
        db.execute(text("DELETE FROM usuarios WHERE empresa_id = :emp_id"), {"emp_id": emp_id_int})
        db.execute(text(f"DELETE FROM {tabla_real} WHERE id = :id"), {"id": empresa_id})
        db.commit()
        flash("Se ha eliminado la organización de la base de datos.", "success")
    except Exception as e:
        db.rollback()
        flash(f"❌ Error al eliminar organización: {str(e)}", "danger")
    finally:
        db.close()
    return redirect('/admin')

@app.route('/admin/conmutar-estado/<int:empresa_id>')
def conmutar_estado(empresa_id):
    if 'usuario_id' not in session or session['rol'] not in ['admin', 'coordinador']: return redirect('/login')
    db = SessionLocal()
    tabla_real = Empresa.__table__.name
    org = db.execute(text(f"SELECT cerrada, nombre FROM {tabla_real} WHERE id = :id"), {"id": empresa_id}).fetchone()
    if org:
        nuevo_estado = 1 if int(org[0] or 0) == 0 else 0
        db.execute(text(f"UPDATE {tabla_real} SET cerrada = :est WHERE id = :id"), {"est": nuevo_estado, "id": empresa_id})
        db.commit()
    db.close()
    return redirect('/admin')

# =================================================================
# 📥 PASO 2: CARGAR BASE DE COLABORADORES (ESTRUCTURA RÍGIDA A, B, C)
# =================================================================
@app.route('/admin/cargar-colaboradores', methods=['POST'])
def cargar_colaboradores():
    if 'usuario_id' not in session or session['rol'] not in ['admin', 'coordinador']: 
        return redirect('/login')
        
    empresa_id_raw = request.form.get('empresa_id')
    archivo = request.files.get('archivo_colaboradores')
    
    if not empresa_id_raw or not archivo:
        flash("Faltan campos requeridos para cargar la base de colaboradores.", "danger")
        return redirect('/admin')
        
    db = SessionLocal()
    try:
        empresa_id = int(empresa_id_raw)
        nombre_archivo = archivo.filename.lower()
        
        # 1. Lectura de Excel o CSV
        if nombre_archivo.endswith('.csv'):
            try:
                df = pd.read_csv(archivo, encoding='utf-8')
            except Exception:
                archivo.seek(0)
                df = pd.read_csv(archivo, encoding='latin1')
        else:
            df = pd.read_excel(archivo)
            
        df = df.where(pd.notnull(df), None)
        columnas_originales = df.columns.tolist()
        
        # 2. ESTRUCTURA RÍGIDA ESTÁNDAR (Pila fija de 3 columnas iniciales)
        col_id_name = columnas_originales[0]      # Columna A: Identificación
        col_nombre_name = columnas_originales[1]  # Columna B: Nombre
        col_email_name = columnas_originales[2]   # Columna C: E-Mail
        
        # Columna D en adelante: Demografías analíticas
        columnas_demograficas = columnas_originales[3:]

        # 3. Limpieza de datos previos en Neon para esta empresa
        colab_ids = [c[0] for c in db.query(Colaborador.id).filter(Colaborador.empresa_id == empresa_id).all()]
        if colab_ids:
            db.query(ValorDemografico).filter(ValorDemografico.colaborador_id.in_(colab_ids)).delete(synchronize_session=False)
            db.query(Participacion).filter(Participacion.colaborador_id.in_(colab_ids)).delete(synchronize_session=False)
            db.query(Colaborador).filter(Colaborador.empresa_id == empresa_id).delete(synchronize_session=False)
            
        db.query(CategoriaDemografica).filter(CategoriaDemografica.empresa_id == empresa_id).delete(synchronize_session=False)
        db.commit()

        # 4. Inserción masiva de Categorías Demográficas (Columna D+)
        categorias_dicts = [{'nombre': str(col_demog).strip(), 'empresa_id': empresa_id} for col_demog in columnas_demograficas]
        db.bulk_insert_mappings(CategoriaDemografica, categorias_dicts)
        db.commit()
        
        cats_db = db.query(CategoriaDemografica).filter(CategoriaDemografica.empresa_id == empresa_id).all()
        mapa_categorias = {cat.nombre: cat.id for cat in cats_db}

        # 5. Inserción masiva de Colaboradores (Lectura A, B, C)
        colaboradores_dicts = []
        for index, fila in df.iterrows():
            id_raw = fila[col_id_name]
            nombre = fila[col_nombre_name]
            email = fila[col_email_name]
            
            if id_raw is None or email is None:
                continue
                
            identificacion = str(id_raw).strip().split('.')[0]
            email_limpio = str(email).strip().lower()
            
            colaboradores_dicts.append({
                'identificacion': identificacion,
                'nombre': str(nombre).strip(),
                'email': email_limpio,
                'empresa_id': empresa_id
            })
            
        db.bulk_insert_mappings(Colaborador, colaboradores_dicts)
        db.commit()

        # 6. Inserción masiva de Valores Demográficos (Cruza por Columna C / E-Mail)
        colabs_db = db.query(Colaborador.id, Colaborador.email).filter(Colaborador.empresa_id == empresa_id).all()
        mapa_colaboradores_by_email = {c.email.strip().lower(): c.id for c in colabs_db if c.email}
        
        valores_dicts = []
        for index, fila in df.iterrows():
            email_raw = fila[col_email_name]
            if email_raw is None:
                continue
            email_val = str(email_raw).strip().lower()
            colab_id = mapa_colaboradores_by_email.get(email_val)
            
            if not colab_id:
                continue
                
            for col_demog in columnas_demograficas:
                valor_celda = fila[col_demog]
                if valor_celda is not None:
                    valor_str = str(valor_celda).strip()
                    if valor_str and valor_str.lower() != 'nan':
                        valores_dicts.append({
                            'colaborador_id': colab_id,
                            'categoria_id': mapa_categorias[str(col_demog).strip()],
                            'valor': valor_str[:150]
                        })
                        
        db.bulk_insert_mappings(ValorDemografico, valores_dicts)
        db.commit()

        flash(f"👥 Censo cargado exitosamente. Se registraron {len(colaboradores_dicts)} colaboradores y {len(valores_dicts)} segmentos demográficos.", "success")
        
    except Exception as e:
        db.rollback()
        print(f"❌ Error al cargar colaboradores: {e}")
        flash(f"Error al procesar la base de colaboradores: {str(e)}", "danger")
    finally:
        db.close()
        
    return redirect('/admin')
# =================================================================
# 📊 PASO 3: CARGAR PARTICIPACIÓN (LEE COLUMNA A)
# =================================================================
@app.route('/admin/cargar-participacion', methods=['POST'])
def cargar_participacion():
    if 'usuario_id' not in session or session['rol'] not in ['admin', 'coordinador']: 
        return redirect('/login')
        
    empresa_id = request.form.get('empresa_id')
    archivo = request.files.get('archivo_participacion')
    
    if not empresa_id or not archivo: 
        flash("Por favor seleccione la organización y el archivo de participación.", "warning")
        return redirect('/admin')
        
    db = SessionLocal()
    try:
        empresa_id_int = int(empresa_id)
        nombre_archivo = archivo.filename.lower()
        
        if nombre_archivo.endswith('.csv'):
            try:
                df = pd.read_csv(archivo, encoding='utf-8')
            except Exception:
                archivo.seek(0)
                df = pd.read_csv(archivo, encoding='latin1')
        else:
            df = pd.read_excel(archivo)
            
        col_email_idx = 0
        
        colaboradores_db = db.query(Colaborador.id, Colaborador.email).filter(Colaborador.empresa_id == empresa_id_int).all()
        mapa_colaboradores = {c.email.strip().lower(): c.id for c in colaboradores_db if c.email}
        
        participaciones_existentes = {p[0] for p in db.query(Participacion.colaborador_id).all()}
        
        nuevas_participaciones = []
        conteo_respuestas = 0
        
        for _, fila in df.iterrows():
            email_raw = fila.iloc[col_email_idx]
            if pd.isna(email_raw) or not email_raw:
                continue
                
            email_val = str(email_raw).strip().lower()
            
            if email_val in mapa_colaboradores:
                colab_id = mapa_colaboradores[email_val]
                if colab_id not in participaciones_existentes:
                    nuevas_participaciones.append({'colaborador_id': colab_id, 'contesto': True})
                    participaciones_existentes.add(colab_id)
                    conteo_respuestas += 1
                    
        if nuevas_participaciones:
            db.bulk_insert_mappings(Participacion, nuevas_participaciones)
            db.commit()
            flash(f"✅ Reporte de participación actualizado. Se cruzaron y registraron {conteo_respuestas} respuestas recibidas.", "success")
        else:
            flash("ℹ️ Las participaciones de este archivo ya se encontraban registradas previamente.", "info")
            
    except Exception as e:
        db.rollback()
        print(f"❌ Error en carga de participación: {str(e)}")
        flash(f"⚠️ Error al procesar el archivo de participación: {str(e)}", "danger")
    finally:
        db.close()
        
    return redirect('/admin')

@app.route('/restablecer-password', methods=['POST'])
def restablecer_password():
    identificador = request.form.get('identificador', '').strip().lower()
    if not identificador: 
        return redirect('/login')
        
    db = SessionLocal()
    tabla_real = Empresa.__table__.name
    usuario, empresa_asoc = None, None
    
    usuario = db.query(Usuario).filter(Usuario.email == identificador).first()
    if usuario and usuario.empresa_id:
        emp_res = db.execute(text(f"SELECT id, nombre, nombre_cliente, correo_cliente, correo_coordinador FROM {tabla_real} WHERE id = :id"), {"id": usuario.empresa_id}).fetchone()
        if emp_res:
            empresa_asoc = emp_res
            
    if usuario:
        nueva_clave = generar_password_aleatorio()
        usuario.password_hash = nueva_clave
        db.commit()
        
        destinatario = (empresa_asoc[3] if empresa_asoc and empresa_asoc[3] else None) or (empresa_asoc[4] if empresa_asoc and len(empresa_asoc) > 4 else usuario.email)
        nombre_dest = (empresa_asoc[2] if empresa_asoc and empresa_asoc[2] else None) or (empresa_asoc[1] if empresa_asoc else 'Usuario')
        
        try:
            enviar_correo_notificacion(destinatario, nombre_dest, usuario.email, nueva_clave, es_olvido=True)
        except Exception as e:
            print(f"❌ Error al enviar mail de clave: {e}")
            
        flash("🔑 Nueva contraseña generada y notificada.", "success")
    else:
        flash("⚠️ No se encontró ningún usuario con esa identificación/correo.", "danger")
        
    db.close()
    return redirect('/login')

def calcular_margen_error(esperadas, recibidas):
    """
    Fórmula exacta de GCTI:
    =IF(C2>B2, "-", IF(C2="", "", 0.0196*50*(B2-C2)/(SQRT(C2)*(B2-1))*100))
    """
    try:
        b2 = float(esperadas)
        c2 = float(recibidas)

        if c2 > b2:
            return "-"
        if c2 == 0 or b2 <= 1:
            return 0.0

        numerador = 0.0196 * 50 * (b2 - c2)
        denominador = math.sqrt(c2) * (b2 - 1)

        resultado = (numerador / denominador) * 100
        return round(resultado, 2)
    except Exception:
        return "-"

@app.route('/api/metricas/<int:categoria_id>')
def obtener_metricas(categoria_id):
    if 'usuario_id' not in session: return jsonify({"error": "No autorizado"}), 401
    db = SessionLocal()
    categoria = db.query(CategoriaDemografica).filter(CategoriaDemografica.id == categoria_id).first()
    if not categoria:
        db.close()
        return jsonify({"error": "Categoría no encontrada"}), 404
        
    if session['rol'] == 'cliente' and categoria.empresa_id != session['empresa_id']:
        db.close()
        return jsonify({"error": "No autorizado"}), 403
        
    query_resultados = db.query(
        ValorDemografico.valor,
        func.count(Colaborador.id).label('total'),
        func.count(Participacion.id).label('contestaron')
    ).join(Colaborador, ValorDemografico.colaborador_id == Colaborador.id)\
    .outerjoin(Participacion, Colaborador.id == Participacion.colaborador_id)\
    .filter(ValorDemografico.categoria_id == categoria_id, Colaborador.empresa_id == categoria.empresa_id)\
    .group_by(ValorDemografico.valor).all()
    db.close()
    
    data_json = []
    for fila in query_resultados:
        b2 = float(fila.total)
        c2 = float(fila.contestaron)
        
        if c2 > b2:
            margen = "-"
        elif c2 == 0 or b2 <= 1:
            margen = 0.0
        else:
            try:
                num = 0.0196 * 50 * (b2 - c2)
                den = math.sqrt(c2) * (b2 - 1)
                margen = round((num / den) * 100, 2)
            except Exception:
                margen = "-"

        if 0 < fila.contestaron < 5:
            data_json.append({
                "opcion": fila.valor, 
                "total_colaboradores": fila.total, 
                "han_contestado": "-", 
                "porcentaje_participacion": "-", 
                "margen_error": "-",
                "anonimo": True
            })
        else:
            pct = round((c2 / b2) * 100, 1) if b2 > 0 else 0
            data_json.append({
                "opcion": fila.valor, 
                "total_colaboradores": fila.total, 
                "han_contestado": fila.contestaron, 
                "porcentaje_participacion": pct, 
                "margen_error": margen,
                "anonimo": False
            })
    return jsonify(data_json)

@app.route('/admin/visor-reportes')
def visor_reportes_admin():
    if 'usuario_id' not in session or session['rol'] not in ['admin', 'coordinador']: return redirect('/login')
    db = SessionLocal()
    empresas = db.query(Empresa).all()
    empresa_seleccionada_id = request.args.get('empresa_id')
    categorias = []
    if empresa_seleccionada_id:
        categorias = db.query(CategoriaDemografica).filter(CategoriaDemografica.empresa_id == empresa_seleccionada_id).all()
    db.close()
    return render_template('admin_reportes.html', empresas=empresas, categories=categorias, categorias=categorias, empresa_sel_id=empresa_seleccionada_id)

@app.route('/api/progreso-global')
def progreso_global():
    empresa_id = request.args.get('empresa_id')
    if not empresa_id:
        return jsonify({"total_enviadas": 0, "total_respondidas": 0, "porcentaje": 0.0})
        
    db = SessionLocal()
    try:
        emp_id = int(empresa_id)
        total_enviadas = db.query(func.count(Colaborador.id)).filter(Colaborador.empresa_id == emp_id).scalar() or 0
        total_respondidas = db.query(func.count(Participacion.id)).join(Colaborador).filter(Colaborador.empresa_id == emp_id).scalar() or 0
        porcentaje = (total_respondidas / total_enviadas * 100) if total_enviadas > 0 else 0.0
        
        return jsonify({
            "total_enviadas": total_enviadas,
            "total_respondidas": total_respondidas,
            "porcentaje": round(porcentaje, 1)
        })
    except Exception as e:
        print(f"❌ Error en progreso global: {e}")
        return jsonify({"total_enviadas": 0, "total_respondidas": 0, "porcentaje": 0.0})
    finally:
        db.close()

def generar_excel_multihoja_gcti(empresa_id, categorias_ids_seleccionadas):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    db = SessionLocal()
    
    try:
        empresa = db.query(Empresa).filter(Empresa.id == int(empresa_id)).first()
        nombre_empresa = empresa.nombre if empresa else 'Organización'

        # Fecha actual formateada y título unificado en A1:D3
        meses_es = ["enero", "febrero", "marzo", "abril", "mayo", "junio", "julio", "agosto", "septiembre", "octubre", "noviembre", "diciembre"]
        hoy = datetime.now()
        fecha_formateada = f"{hoy.day:02d} de {meses_es[hoy.month - 1]} de {hoy.year}"
        titulo_encabezado = f"Great Culture to Innovate México - Reporte de Participación {fecha_formateada}\n{nombre_empresa}"

        COLORES_NIVELES = {
            1: 'C00000', 2: '0000FF', 3: '70AD47', 4: 'C65911', 5: '7030A0', 6: 'A6A6A6', 7: '000000'
        }

        fill_encabezado = PatternFill(start_color='000000', end_color='000000', fill_type='solid')
        font_encabezado = Font(name='Century Gothic', size=10, bold=True, color='FFFFFF')
        alineacion_centro = Alignment(horizontal='center', vertical='center', wrap_text=True)
        alineacion_izquierda = Alignment(horizontal='left', vertical='center')
        
        borde_fino = Side(border_style='thin', color='D9D9D9')
        borde_celda = Border(left=borde_fino, right=borde_fino, top=borde_fino, bottom=borde_fino)

        path_logo_gcti = os.path.join(app.root_path, 'static', 'logo_gcti.png')

        # Normalizar IDs de categorías
        ids_limpios = []
        for cid in categorias_ids_seleccionadas:
            try:
                ids_limpios.append(int(cid))
            except (ValueError, TypeError):
                continue

        grupos_categorias = {}
        for cat_id_int in ids_limpios:
            cat_obj = db.query(CategoriaDemografica).filter(
                CategoriaDemografica.id == cat_id_int,
                CategoriaDemografica.empresa_id == int(empresa_id)
            ).first()

            if not cat_obj:
                continue

            nombre_base = re.sub(r'\s+\d+$', '', cat_obj.nombre).strip()
            match_nivel = re.search(r'(\d+)$', cat_obj.nombre)
            nivel_num = int(match_nivel.group(1)) if match_nivel else 1

            if nombre_base not in grupos_categorias:
                grupos_categorias[nombre_base] = []

            grupos_categorias[nombre_base].append({'id': cat_obj.id, 'nombre_original': cat_obj.nombre, 'nivel': nivel_num})

        colabs_db = db.query(Colaborador.id).filter(Colaborador.empresa_id == int(empresa_id)).all()
        colab_ids_list = [c[0] for c in colabs_db]
        participaciones_set = {p[0] for p in db.query(Participacion.colaborador_id).filter(Participacion.colaborador_id.in_(colab_ids_list)).all()} if colab_ids_list else set()

        for nombre_grupo, lista_subcats in grupos_categorias.items():
            lista_subcats = sorted(lista_subcats, key=lambda x: x['nivel'])
            nombre_hoja = str(nombre_grupo).replace('/', '-').replace('\\', '-')[:30]
            ws = wb.create_sheet(title=nombre_hoja)

            # 1. Quitar cuadrícula gris del fondo
            ws.views.sheetView[0].showGridLines = False

            # 2. Definir anchos fijos de columnas exactos
            ws.column_dimensions['A'].width = 80
            ws.column_dimensions['B'].width = 25
            ws.column_dimensions['C'].width = 25
            ws.column_dimensions['D'].width = 25
            ws.column_dimensions['E'].width = 25

            # 3. Alturas de cabecera
            ws.row_dimensions[1].height = 20
            ws.row_dimensions[2].height = 20
            ws.row_dimensions[3].height = 20
            ws.row_dimensions[4].height = 12  # Fila vacía

            # 4. COMBINAR Y CONFIGURAR A1:D3 Y E1:E3
            ws.merge_cells('A1:D3')
            ws.merge_cells('E1:E3')

            cell_title = ws['A1']
            cell_title.value = titulo_encabezado
            cell_title.font = Font(name='Century Gothic', size=11, bold=True, color='000000')
            cell_title.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

            # Estampar Logo GCTI en E1:E3
            if os.path.exists(path_logo_gcti):
                try:
                    img_gcti = openpyxl_image.Image(path_logo_gcti)
                    img_gcti.width = 90
                    img_gcti.height = 45
                    ws.add_image(img_gcti, 'E1')
                except Exception as e:
                    print(f"⚠️ Alerta Logo GCTI: {e}")

            # 5. FILA 5: ENCABEZADOS DE TABLA (FRANJA NEGRA)
            ws.row_dimensions[5].height = 24
            headers = [nombre_grupo, 'Población objetivo', 'Encuestas recibidas', '(%) avance', 'Margen de error (%)']

            for col_idx, text_header in enumerate(headers, 1):
                cell = ws.cell(row=5, column=col_idx, value=text_header)
                cell.fill = fill_encabezado
                cell.font = font_encabezado
                cell.alignment = alineacion_centro if col_idx > 1 else alineacion_izquierda

            # Cargar valores demográficos
            valores_por_colab = {}
            cats_ids_grupo = [s['id'] for s in lista_subcats]
            
            if colab_ids_list:
                registros_valores = db.query(ValorDemografico).filter(
                    ValorDemografico.colaborador_id.in_(colab_ids_list),
                    ValorDemografico.categoria_id.in_(cats_ids_grupo)
                ).all()
                for rv in registros_valores:
                    if rv.colaborador_id not in valores_por_colab:
                        valores_por_colab[rv.colaborador_id] = {}
                    valores_por_colab[rv.colaborador_id][rv.categoria_id] = rv.valor

            # 6. FILAS 6+: POBLADO DE DATOS Y ÁRBOL JERÁRQUICO
            def procesar_nivel_recursivo(colabs_subconjunto, subcat_idx):
                if subcat_idx >= len(lista_subcats) or not colabs_subconjunto:
                    return

                subcat_actual = lista_subcats[subcat_idx]
                cat_id_curr = subcat_actual['id']
                nivel_curr = subcat_actual['nivel']

                agrupados = {}
                for colab_id in colabs_subconjunto:
                    val = valores_por_colab.get(colab_id, {}).get(cat_id_curr)
                    if val and str(val).strip() and str(val).lower() != 'nan':
                        val_str = str(val).strip()
                        if val_str not in agrupados:
                            agrupados[val_str] = []
                        agrupados[val_str].append(colab_id)

                for val_nombre, ids_hijos in agrupados.items():
                    b2 = float(len(ids_hijos))
                    c2 = float(sum(1 for cid in ids_hijos if cid in participaciones_set))

                    if 0 < c2 < 5:
                        row_data = [val_nombre, int(b2), '-', '-', '-']
                    else:
                        pct = round((c2 / b2) * 100, 1) if b2 > 0 else 0.0
                        margen = '-' if c2 > b2 else (0.0 if (c2 == 0 or b2 <= 1) else calcular_margen_error(b2, c2))
                        margen_str = f'{margen}%' if margen != '-' else '-'
                        pct_str = f'{pct}%'
                        row_data = [val_nombre, int(b2), int(c2), pct_str, margen_str]

                    ws.append(row_data)
                    row_idx = ws.max_row
                    ws.row_dimensions[row_idx].height = 20

                    color_hex = COLORES_NIVELES.get(nivel_curr, '000000')
                    font_nivel = Font(name='Century Gothic', size=10, bold=(nivel_curr <= 3), color=color_hex)
                    alineacion_sangria = Alignment(horizontal='left', vertical='center', indent=max(0, nivel_curr - 1))

                    for col_idx in range(1, 6):
                        cell = ws.cell(row=row_idx, column=col_idx)
                        cell.border = borde_celda
                        if col_idx == 1:
                            cell.font = font_nivel
                            cell.alignment = alineacion_sangria
                        else:
                            cell.font = Font(name='Century Gothic', size=10)
                            cell.alignment = alineacion_centro

                    procesar_nivel_recursivo(ids_hijos, subcat_idx + 1)

            procesar_nivel_recursivo(colab_ids_list, 0)

    finally:
        db.close()

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output
    
@app.route('/admin/descargar-reporte-excel', methods=['POST'])
def descargar_reporte_excel():
    if 'usuario_id' not in session or session['rol'] not in ['admin', 'coordinador']:
        return jsonify({'error': 'No autorizado'}), 401

    try:
        data = request.get_json() or {}
        empresa_id_raw = data.get('empresa_id')
        categorias_ids = data.get('categorias_ids', [])

        if not empresa_id_raw or not categorias_ids:
            return jsonify({'error': 'Debe seleccionar la organización y al menos una demografía.'}), 400

        empresa_id = int(empresa_id_raw)
        db = SessionLocal()
        empresa = db.query(Empresa).filter(Empresa.id == empresa_id).first()
        nombre_empresa = empresa.nombre if empresa else 'Organizacion'
        db.close()

        excel_stream = generar_excel_multihoja_gcti(empresa_id, categorias_ids)
        excel_stream.seek(0)
        nombre_archivo = f"Reporte_GCTI_{nombre_empresa.replace(' ', '_')}.xlsx"

        return send_file(
            excel_stream,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=nombre_archivo
        )
    except Exception as e:
        print(f"❌ Error crítico al generar Excel: {str(e)}")
        return jsonify({'error': f'Falló la generación del archivo Excel: {str(e)}'}), 500

@app.route('/admin/enviar-reporte-email', methods=['POST'])
def enviar_reporte_email():
    if 'usuario_id' not in session or session['rol'] not in ['admin', 'coordinador']:
        return jsonify({'error': 'No autorizado'}), 401

    data = request.get_json() or {}
    empresa_id = data.get('empresa_id')
    categorias_ids = data.get('categorias_ids', [])

    if not empresa_id or not categorias_ids:
        return jsonify({'error': 'Debe seleccionar al menos una demografía.'}), 400

    db = SessionLocal()
    empresa = db.query(Empresa).filter(Empresa.id == empresa_id).first()
    nombre_empresa = empresa.nombre if empresa else 'Organización'
    db.close()

    excel_stream = generar_excel_multihoja_gcti(empresa_id, categorias_ids)
    excel_stream.seek(0)

    remitente_autenticado = 'carlos.mora@peoplesvoice.co'
    nombre_remitente = session.get('nombre', 'Portal GCTI®')

    destinatario_principal = 'roberto.cruz@greatculturetoinnovate.net'
    copia_cc = 'german.romero@peoplesvoice.co'

    msg = MIMEMultipart()
    msg['From'] = f"Portal de Reportes GCTI® <{remitente_autenticado}>"
    msg['To'] = destinatario_principal
    msg['Cc'] = copia_cc
    msg['Subject'] = f'Envío de reporte de participación — {nombre_empresa}'

    cuerpo_html = f"""
    <html>
    <body style="font-family: 'Helvetica Neue', Arial, sans-serif; color: #333; line-height: 1.6;">
        <div style="max-width: 600px; margin: 0 auto; padding: 25px; border: 1px solid #edf2f7; border-radius: 10px; background-color: #ffffff;">
            <p>Buen día,</p>
            <p>Adjunto el reporte de participación correspondiente a la medición en curso para la organización <strong>{nombre_empresa}</strong>, para su revisión.</p>
            <p>El documento consolida los principales indicadores de participación a la fecha. Quedo atento a sus comentarios o a cualquier información adicional que requieran.</p>
            <br>
            <p>Cordialmente,</p>
            <p><strong>{nombre_remitente}</strong></p>
            <hr style="border: 0; border-top: 1px solid #edf2f7; margin: 25px 0 15px 0;">
            <p style="font-size: 0.8rem; color: #a0aec0; text-align: center;">Great Culture to Innovate® — Peoples Voice © 2026</p>
        </div>
    </body>
    </html>
    """
    msg.attach(MIMEText(cuerpo_html, 'html'))

    parte_adjunto = MIMEApplication(
        excel_stream.read(),
        _subtype='vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    nombre_archivo = f"Reporte_GCTI_{nombre_empresa.replace(' ', '_')}.xlsx"
    parte_adjunto.add_header('Content-Disposition', 'attachment', filename=nombre_archivo)
    msg.attach(parte_adjunto)

    try:
        # CAMBIO CLAVE: Usar SMTP_SSL en puerto 465 (sin bloqueo de red en Render)
        server = smtplib.SMTP_SSL('smtp.gmail.com', 465)
        server.login(remitente_autenticado, 'hxhjkhqleflgvmoo')
        recipients = [destinatario_principal, copia_cc]
        server.sendmail(remitente_autenticado, recipients, msg.as_string())
        server.quit()
        return jsonify({'success': True, 'mensaje': f'📧 Reporte de {nombre_empresa} enviado exitosamente por correo.'})
    except Exception as e:
        print(f"❌ Error SMTP: {str(e)}")
        return jsonify({'error': f'Falló el envío por correo electrónico: {str(e)}'}), 500
        
print("5. Intentando encender el servidor Flask en el entorno de Render...")

if __name__ == '__main__':
    puerto = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=puerto)
